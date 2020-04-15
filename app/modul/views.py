from flask import render_template, request, session, redirect,url_for
from app import app
from openpyxl import load_workbook
import os
import pandas as pd
import dbmodel as x
from nltk.tokenize import word_tokenize
from Sastrawi.StopWordRemover.StopWordRemoverFactory import StopWordRemoverFactory
remover = StopWordRemoverFactory().create_stop_word_remover()
from Sastrawi.Stemmer.StemmerFactory import StemmerFactory
stemmer = StemmerFactory().create_stemmer()
import numpy as np
import math
from sklearn.metrics.pairwise import cosine_similarity
#from sklearn.metrics.pairwise import euclidean_distances
import string
from flask import jsonify
from sklearn.feature_extraction.text import CountVectorizer
from sklearn.feature_extraction.text import TfidfVectorizer
import numpy as np
import math
from feature_extraction import FeatureExtraction
import json
import plotly


def getRandom(k, data_len):
    a = np.arange(data_len)
    np.random.shuffle(a)
    return a[:k]


# getRandom(2,9)
def getLowest(d):
    d = np.array(d)
    tempSort = d.argsort()[:4]
    sortedlist = tempSort[1:]
    return sortedlist


def calcCenOne(d, nCen):
    #     print nCen
    d1 = pow(d[nCen], 3)
    d2 = 0.0

    for val in d:
        if val != 0.0:
            d2 = d2 + 1 / pow(val, 2)

    d2 = pow(d2, 2)
    if d1 == 0.0 or d2 == 0.0:
        return 0.0
    else:
        return 1 / (d1 * d2)
    print nCen


def getNearest(array, values):
    hasil = []
    idx = (np.abs(array - values)).argmin()

    return array[idx]


def find_nearest(array, values):
    hasil = []
    idx = []
    a = np.asarray(array)
    for val in values:
        tempval = getNearest(a, val)
        if tempval in hasil:
            while tempval in hasil:
                index = np.argwhere(a == tempval)
                a = np.delete(a, index)
                tempval = getNearest(a, val)
            hasil.append(getNearest(a, val))
        else:
            hasil.append(tempval)
    return hasil


def kHarmonic(score_ec, k, data_len, total_tf):
    id_cluster = getRandom(k, data_len)
    #     id_cluster = [1,8]
    iteration = 0
    print id_cluster

    clusterbefore = []
    for i in range(0, k):
        clusterbefore.append([])

    while True:

        iteration = iteration + 1
        hasil = []
        centroid = []
        print iteration

        for i in range(0, k):
            string_cluster = "cluster" + str(i) + " = []"
            exec (string_cluster)
            #             print string_cluster
            string_pembilang = "pembilang" + str(i) + " = 0.0"
            exec (string_pembilang)
            #             print string_pembilang
            string_penyebut = "penyebut" + str(i) + " = 0.0"
            exec (string_penyebut)
        #             print string_penyebut

        for row in range(0, data_len):
            temp_cluster = []
            for cluster in id_cluster:
                temp_cluster.append((score_ec[row][cluster]))
            #                 print temp_cluster

            for i in range(0, k):
                valbilang = (calcCenOne(temp_cluster, i) * total_tf[row])
                #                 print valbilang
                valsebut = (calcCenOne(temp_cluster, i))
                #                 print valsebut
                #                 print
                string_pembilang1 = "pembilang" + str(i) + " = pembilang" + str(i) + " + " + str(valbilang)
                exec (string_pembilang1)
                #                 print string_pembilang1
                string_penyebut1 = "penyebut" + str(i) + " = penyebut" + str(i) + " + " + str(valsebut)
                exec (string_penyebut1)
            #                 print string_penyebut1

            index_cluster = temp_cluster.index(min(temp_cluster))
            string_append_cluster = "cluster" + str(index_cluster) + ".append(" + str(row) + ")"
            locals()['cluster' + str(index_cluster)].append(row)
        #             print index_cluster
        #             print string_append_cluster

        for i in range(0, k):
            hasil.append(locals()['cluster' + str(i)])

        for i in range(0, k):
            string_centroid1 = "centroid.append(pembilang" + str(i) + "/penyebut" + str(i) + ")"
            exec (string_centroid1)

        nearidx = []
        print "centroid : " + str(centroid)
        nearval = find_nearest(total_tf, centroid)
        for nval in nearval:
            nearidx.append(total_tf.index(nval))

        id_cluster = nearidx

        print "before" + str(clusterbefore)
        print "after" + str(hasil)

        status_cluster = 1
        for y in range(0, k):
            if hasil[y] in clusterbefore:
                status_cluster = status_cluster + 1
        print status_cluster
        if status_cluster == (k + 1):
            break
        # else:
        #     if iteration == max_i:
        #         id_cluster = getRandom(k, data_len)
        #         iteration = 0
        clusterbefore = hasil
        print centroid
        print id_cluster

    return hasil

@app.route('/')
def form_lo():
    if 'username' in session:
        return render_template('index.html')

    return render_template("login.html")

@app.route('/login', methods = ["POST"])
def login():
    user = request.form['username']
    dbmodel = x.DBModel()
    login_user = dbmodel.find_user("Judul_Skripsi","user", user)

    if login_user:
        if request.form['pass'].encode('utf-8')== login_user['pass'].encode('utf-8'):
            session['username'] = request.form['username']
            return redirect(url_for('form_lo'))

    return render_template("login.html")


@app.route('/logout')
def logout():
    session.pop('username',None)
    return render_template("login.html")

@app.route('/index')
def index():
    # if 'username' not in session:
    #     return render_template('login.html')

    return render_template("index.html")

@app.route('/uploading')
def uploading():
    # if 'username' not in session:
    #     return render_template('login.html')

    return render_template("upload_data.html")

@app.route('/hasilupload', methods = ['GET', 'POST'])
def hasilupload():
    # if 'username' not in session:
    #     return render_template('login.html')

    if request.method == 'POST':
        f = request.files['file']
        file = f.filename
        s = request.form['sheet']
        f.save(os.path.join('app/upload_data', file))

        dbmodel = x.DBModel()
        find_file = dbmodel.find_file("Judul_Skripsi", "file", file, s)

        if find_file == True:
            dbmodel.delete_same("Judul_Skripsi", "file", file, s)

        dbmodel.insert_file("Judul_Skripsi", "file", file, s)

        get_file = dbmodel.get_file_desc("Judul_Skripsi", "file")
        for w in get_file:
            values = w.values()
            for y in values:
                y = y

        wb = load_workbook(filename="app/upload_data/"+file)
        for u in wb.get_sheet_names():
            if u == y:
                sheet_available = True
                break
            else:
                sheet_available = False

        if sheet_available == False:
            return render_template("upload_data.html")

        sheet_ranges = wb[request.form['sheet']]
        data = pd.DataFrame(sheet_ranges.values)

    return render_template('hasilexcel.html', tables=[data.to_html(classes='table table-striped table-bordered table-hover')])

@app.route('/hasilpemilihan', methods = ['GET', 'POST'])
def hasilpemilihan():
    # if 'username' not in session:
    #     return render_template('login.html')

    if request.method == 'POST':

        select1 = request.form["select1"]
        select2 = request.form["select2"]
        selectcolom =request.form["selectcolom"]
        namacolom = request.form["namacolom"]

        dbmodel = x.DBModel()
        # get sheet dari database, hasil upload terakhir
        get_sfile = dbmodel.get_file_desc("Judul_Skripsi", "file")
        for w in get_sfile:
            values = w.values()
            for y in values:
                sheet = y

        # get nama file dari database, hasil upload terakhir
        get_nfile = dbmodel.get_file_desc2("Judul_Skripsi", "file")
        for n in get_nfile:
            values = n.values()
            for q in values:
                file = q

        wb = load_workbook(filename='app/upload_data/'+file)
        sheet_ranges = wb[sheet]
        data = pd.DataFrame(sheet_ranges.values)

        row1 = int(select1)
        row2 = int(select2)

        cols = selectcolom.split(",") #memisahkan inputan kolom dipilih berdasarkan koma
        cols = list(map(int,cols)) #corvert to int
        xname = namacolom.split(",") #memisahkan inptan nama kolom berdasarkan koma
        data =data[row1:row2][cols]#data terpilih berdasarkan inputan baris dan kolom
        data.columns = [xname]
        data = data.dropna()

        header = {}
        for index, head in enumerate(xname):
            header[str(index)] = head

        pd.options.display.max_colwidth = 999
        print (data)
        data.encode("utf-8")  # To utf-8 encoding scheme
        data.encode('ascii', 'ignore')
        print (data)

        encoded_string = map(str, data)
        print (encoded_string)

        dbmodel = x.DBModel() #memanggil file model dimodel class DBModel
        result_insert_table= dbmodel.insert_cleaning_data("Judul_Skripsi","datanya",encoded_string)
        result_insert_header = dbmodel.insert_header("Judul_Skripsi","judulnya",header)

    return render_template('masukprosessing.html', tables=[data.to_html(classes='table table-striped table-bordered table-hover')])

@app.route('/tokenisasi', methods = ['GET', 'POST'])
def tokenisasi():
    # if 'username' not in session:
    #     return render_template('login.html')

    dbmodel = x.DBModel()
    get_data =dbmodel.get_data_all("Judul_Skripsi","datanya")

    data_all =[]
    for i in get_data:
        get_values = i.values()
        a = get_values
        for u in a:
            data_lower = u.lower()
            word_token = word_tokenize(data_lower)
        data_all.append(word_token)

    data = pd.DataFrame(data_all)
    head_token = []
    for index in data.columns:
        custom_head = "K" + str(index)
        head_token.append(custom_head)
    data.columns = head_token

    dbmodel = x.DBModel()  # memanggil file model dimodel class DBModel
    dbmodel.insert_tokenisasi("Judul_Skripsi", "Tokenisasi", data)


    return render_template("tokenisasi.html", tables=[data.to_html(classes='table table-striped table-bordered table-hover')])


@app.route('/filtering', methods=['GET', 'POST'])
def filtering():
    # if 'username' not in session:
    #     return render_template('login.html')

    dbmodel = x.DBModel()
    get_data = dbmodel.get_data_all("Judul_Skripsi", "Tokenisasi")

    data_filtering = []
    for i in get_data:
        a = i.values()
        b = a
        fil = []
        for j in b:
            if j <> None:
                data_fil = (j.encode("ascii", "ignore"))
                stopword = remover.remove(data_fil)
                if stopword <> "":
                    fil.append(stopword)
        data_filtering.append(fil)

    data = pd.DataFrame(data_filtering)
    head_filter = []
    for index in data.columns:
        custom_head = "K" + str(index)
        head_filter.append(custom_head)
    data.columns = head_filter

    dbmodel = x.DBModel()  # memanggil file model dimodel class DBModel
    dbmodel.insert_filtering("Judul_Skripsi", "Fitering", data)

    return render_template("filtering.html", tables=[data.to_html(classes='table table-striped table-bordered table-hover')])

@app.route('/stemming', methods=['GET', 'POST'])
def stemming():
    # if 'username' not in session:
    #     return render_template('login.html')

    dbmodel = x.DBModel()
    get_data = dbmodel.get_data_all("Judul_Skripsi","Fitering")

    data_stemming = []
    for i in get_data:
        c = i.values()
        d = c
        data_s = []
        for k in d:
            if k <> None:
                data_stem = (k.encode("ascii", "ignore"))
                stemming = stemmer.stem(data_stem)
                stemming = stemming.translate(None, string.punctuation)
                if stemming <> "":
                    data_s.append(stemming)
        data_stemming.append(data_s)

    data = pd.DataFrame(data_stemming)
    head_stem = []
    for index in data.columns:
        custom_head = "K" + str(index)
        head_stem.append(custom_head)
    data.columns = head_stem

    dbmodel = x.DBModel()  # memanggil file model dimodel class DBModel
    dbmodel.insert_stemming("Judul_Skripsi", "Stemming", data)

    return render_template("stemming.html", tables=[data.to_html(classes='table table-striped table-bordered table-hover')])

@app.route('/inputK', methods = ['GET', 'POST'])
def inputK():
    if 'username' not in session:
        return render_template('login.html')

    return render_template("inputK.html")

@app.route('/harmonic', methods=['GET', 'POST'])
def harmonic():
    if 'username' not in session:
        return render_template('login.html')

    dbmodel = x.DBModel()
    get_stemming = dbmodel.get_data_all("Judul_Skripsi","Stemming")
    get_dataawal = dbmodel.get_data_all("Judul_Skripsi","datanya")

    stem = []
    for h in get_stemming:
        stem1 = []
        val1 = h.values()
        for k in val1:
            if k <> None:
                stem1.append(k)
        stem.append(stem1)
    documents = stem

    awal =   []
    for c in get_dataawal:
        val = c.values()
        for d in val:
            d = d
        awal.append(d)

    # documents = awal

    feature_extraction = FeatureExtraction()
    feature = feature_extraction.fit(documents)
    score_ec = cosine_similarity(feature) #feature)
    #score_ec = euclidean_distances(feature,feature)
    print score_ec


    total_tf = []

    for tf_score in feature:
        total_tf.append(round(np.sum(tf_score), 3))

    k = int(request.form["kluster"])
    data_len = len(feature)

    max_iteration = 100
    lastoutput = kHarmonic(score_ec, k, data_len, total_tf)

    # mengambil nama sheet dari hasil upload terakhir
    get_file = dbmodel.get_file_desc("Judul_Skripsi", "file")
    for w in get_file:
        values = w.values()
        for y in values:
            y = y

    #mengambil nama file dari hasil upload terakhir
    get_file2 = dbmodel.get_file_desc2("Judul_Skripsi", "file")
    for w2 in get_file2:
        values2 = w2.values()
        for y2 in values2:
            y2 = y2

    #mencari sheet didatabase(file) yang nama sheetnya itu = nama sheet yang di uppload terakhir.
    find_sheet = dbmodel.find_sheet("Judul_Skripsi", "file", y)
    #kondisi jika ketemu sheet yang sama maka dihitung ada berapaa sheet
    if find_sheet == True:
        count_sheet = dbmodel.count_sheet("Judul_Skripsi", "file", y)
        y = y + "(" +(str(count_sheet-1)) + ")" # setelah itu nama sheet diubah menjadi (sheet(1))
        dbmodel.update_file1("Judul_Skripsi", "file", y2, y, k) #mengupdate database dengan nama sheet yang baru
        print y

    # update file nambah kolom jumlah kluster
    dbmodel.update_file("Judul_Skripsi", "file", y2, y, k)

    # mencari collection dengan nama sheet dari hasil upload terakhir
    find_collection = dbmodel.find_collection("Judul_Skripsi", y)

    if find_collection == True:
        dbmodel.delete_collection("Judul_Skripsi", y)

    s = 1
    all_out = []
    for out in lastoutput:
        temp_out = []
        for o in out:
            temp_out.append(awal[o])
            dbmodel.insert_hasil("Judul_Skripsi", y, awal[o], s)
        all_out.append(temp_out)
        s = s + 1

    get_hasil = dbmodel.get_data_all("Judul_Skripsi",y)
    table_hasil = pd.DataFrame(list(get_hasil))


    return render_template("k-harmonic.html", tables=[table_hasil.to_html(classes='table table-striped table-bordered table-hover')])

@app.route('/allfile', methods=['GET', 'POST'])
def allfile():
    if 'username' not in session:
        return render_template('login.html')

    dbmodel = x.DBModel()
    get_file = dbmodel.get_data_all("Judul_Skripsi", "file")
    get_sheet = dbmodel.get_sheet("Judul_Skripsi", "file")
    get_nclust = dbmodel.get_nclust("Judul_Skripsi", "file")

    coll_name = []
    for z in get_sheet:
        values = z.values()
        for h in values:
            h = h
            coll_name.append(h)

    nclust = []
    for z2 in get_nclust:
        values2 = z2.values()
        for h2 in values2:
            h2 = str(h2)
            nclust.append(h2)

    table = []
    for k in get_file:
        o = k.values()
        table.append(o)

    return render_template("allfile.html", coll_name=coll_name, tables = table, nclust = nclust)

@app.route('/showresult', methods=['GET', 'POST'])
def showresult():
    if 'username' not in session:
        return render_template('login.html')

    if request.method == 'POST':
        collection = request.form['collection']
        n_clust = request.form['nclust']
        n_clust = int(n_clust)

        dbmodel = x.DBModel()
        get_file = dbmodel.get_data_all("Judul_Skripsi", collection)
        hasil = pd.DataFrame(list(get_file))

        pd.options.display.max_colwidth = 999

        accordion = []
        for p in range(n_clust):
            group = dbmodel.find_group("Judul_Skripsi", collection, p+1)
            accordion.append(list(group))

    return  render_template("file_terpilih.html", tables=[hasil.to_html(classes='table table-striped table-bordered table-hover')], n=n_clust, accordion=accordion)

@app.route('/chart', methods=['GET','POST'])
def chart():
    if 'username' not in session:
        return render_template('login.html')

    # dbmodel = x.DBModel()
    # get_file = dbmodel.get_data_all("Judul_Skripsi", "file")
    #
    # coll_name = []
    # for z in get_file:
    #     values = z.values()
    #     coll_name.append(values)
    #
    # print coll_name
    #
    # all_data = []
    # all_sheet = []
    # for coll in coll_name:
    #     all_sheet.append(coll[1])
    #     temp_cls = []
    #     for p in range(coll[2]):
    #         group = dbmodel.find_group("Judul_Skripsi", coll[1], p+1)
    #         temp_cls.append(len(group))
    #     all_data.append(temp_cls)
    #
    # print all_data
    # print all_sheet
    #
    #
    # datagrafik = []
    #
    # for i in range(0,len(all_data[0])):
    #     tempdict = {}
    #     tempdict["x"] = all_sheet
    #     tempy = []
    #     for dt in all_data:
    #         print dt
    #         tempy.append(dt[i])
    #     tempdict["y"] = tempy
    #     tempdict["type"] = "scatter"
    #     tempdict['mode'] = "lines"
    #     tempdict["name"] = "Cluster" + str(i)
    #     datagrafik.append(tempdict)
    #
    #
    # graphs = [
    #     dict(
    #         data=datagrafik,
    #         layout=dict(
    #             title='Trend Data Skripsi Teknik Informatika'
    #
    #         )
    #     ),
    #
    # ]

    rng = pd.date_range('1/1/2011', periods=7500, freq='H')
    ts = pd.Series(np.random.randn(len(rng)), index=rng)

    graphs = [
        dict(
            data=[
                dict(
                    x=["2012-2013", "2013-2014", "2014-2015", "2015-2016", "2016-2017"],
                    y=[17, 147, 3, 125, 3],
                    type='scatter',
                    mode='lines',
                    name='Aplikasi Mobile dan Pemrograman Web'
                ),
                dict(
                    x=["2012-2013", "2013-2014", "2014-2015", "2015-2016", "2016-2017"],
                    y=[3, 8, 1, 4, 18],
                    type='scatter',
                    mode='lines',
                    name='Sistem Informasi'
                ),
                dict(
                    x=["2012-2013", "2013-2014", "2014-2015", "2015-2016", "2016-2017"],
                    y=[55, 8, 151, 3, 58],
                    type='scatter',
                    mode='lines',
                    name='Multimedia dan Soft Computing'
                )

            ],
            layout=dict(
                title='Trend Data Skripsi Teknik Informatika'

            )
        ),

    ]

    # Add "ids" to each of the graphs to pass up to the client
    # for templating
    ids = ['graph-{}'.format(i) for i, _ in enumerate(graphs)]

    # Convert the figures to JSON
    # PlotlyJSONEncoder appropriately converts pandas, datetime, etc
    # objects to their JSON equivalents
    graphJSON = json.dumps(graphs, cls=plotly.utils.PlotlyJSONEncoder)

    return render_template('grafik.html',
                           ids=ids,
                           graphJSON=graphJSON)

